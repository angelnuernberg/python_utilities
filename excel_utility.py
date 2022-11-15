# Very useful: https://realpython.com/openpyxl-excel-spreadsheets-python/

from openpyxl import load_workbook
workbook = load_workbook(filename="reviews-sample.xlsx")
print(workbook.sheetnames)
sheet=workbook.active
print(sheet.title)
print(f'Value of cell E37: {sheet["E37"].value}')

print('-----------------------------')
print('Editing an existing cell C1 and E5')
sheet["C1"] = "writing ;)"
sheet["E5"] = "writing ;)"
workbook.save(filename="hello_world_append.xlsx")